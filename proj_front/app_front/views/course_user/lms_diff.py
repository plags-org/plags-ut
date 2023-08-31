import dataclasses
import traceback
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
from django.http.request import HttpRequest
from django.http.response import HttpResponse
from django.shortcuts import render

from app_front.core.plags_utils.plags_endpoint import (
    AbsPlagsView,
    annotate_view_endpoint,
)
from app_front.core.storage_util import save_error_text_file
from app_front.dependency.system_notification import SLACK_NOTIFIER
from app_front.forms import UploadCourseUserLmsExcelFile
from app_front.models import Course, CourseUser, Organization, UserAuthorityEnum
from app_front.utils.auth_util import UserAuthorityCapabilityKeys, UserAuthorityDict
from app_front.utils.diff_util import DifferenceDetector, FromToPair
from app_front.utils.exception_util import ExceptionHandler, UserResponsibleException

CommonIDNumberT = str
StudentCardNumberT = str


@dataclasses.dataclass
class StudentInfo:
    common_id_number: CommonIDNumberT
    student_card_number: str
    full_name: str


StudentInfoWithConflict = Tuple[StudentInfo, CourseUser]
StudentInfoWithoutConflict = CourseUser


def _parse_student_card_number(student_card_number: Optional[str]) -> str:
    # NOTE Excelでは空欄がNoneになっている
    # NOTE User.student_card_number は null=False なので空文字列
    if student_card_number is None:
        return ""
    return student_card_number.strip()


class CourseUserLMSDiffView(AbsPlagsView):
    @classmethod
    def _render_page(
        cls,
        request: HttpRequest,
        user_authority: UserAuthorityDict,
        organization: Organization,
        course: Course,
        form: UploadCourseUserLmsExcelFile = None,
    ) -> HttpResponse:
        course_users_only_in_plags: List[CourseUser] = []
        course_users_only_in_lms: List[StudentInfo] = []
        course_users_in_both_with_conflict: List[StudentInfoWithConflict] = []
        course_users_in_both_without_conflict: List[CourseUser] = []
        if form is not None:
            course_user_lms_excel_file = form.cleaned_data["course_user_lms_excel_file"]
            course_user_lms_excel_filename = save_error_text_file(
                course_user_lms_excel_file.read(),
                "course_user_lms_excel_file",
                course_user_lms_excel_file.name,
            )
            try:
                workbook = openpyxl.load_workbook(course_user_lms_excel_filename)
            except Exception as exc:
                error_message = f'Uploaded file "{course_user_lms_excel_file.name}" could not be parsed as Excel file'
                raise UserResponsibleException(error_message) from exc
            assert (
                "課題基本情報" in workbook.sheetnames
            ), 'Uploaded LMS excel workbook has no "課題基本情報" sheet'
            assert (
                "課題全体提出状況" in workbook.sheetnames
            ), 'Uploaded LMS excel workbook has no "課題全体提出状況" sheet'

            worksheet = workbook["課題全体提出状況"]
            assert worksheet.cell(1, 1).value is None, "A1 cell must be empty"
            assert worksheet.cell(1, 2).value is None, "B1 cell must be empty"
            assert (
                worksheet.cell(1, 3).value == "※提出フラグ(0:未提出、1:期限内提出、2.期限外提出、空白:非対象者)"
            ), 'C1 cell must be "※提出フラグ(0:未提出、1:期限内提出、2.期限外提出、空白:非対象者)"'
            assert worksheet.cell(2, 1).value == "ユーザID", 'A2 cell must be "ユーザID"'
            assert worksheet.cell(2, 2).value == "学生証番号", 'B2 cell must be "学生証番号"'
            assert worksheet.cell(2, 3).value == "氏名", 'C2 cell must be "氏名"'

            course_users_in_lms = [
                StudentInfo(
                    common_id_number=common_id_number,
                    student_card_number=_parse_student_card_number(student_card_number),
                    full_name=full_name,
                )
                for (
                    common_id_number,
                    student_card_number,
                    full_name,
                ) in worksheet.iter_rows(min_row=3, max_col=3, values_only=True)
            ]

            course_users: Iterable[CourseUser] = (
                CourseUser.objects.filter(
                    course=course,
                    authority=UserAuthorityEnum(UserAuthorityEnum.STUDENT).value,
                    is_active=True,
                )
                .order_by("-authority", "user__username")
                .select_related(
                    "user", "added_by", "is_active_updated_by", "authority_updated_by"
                )
            )

            course_users_in_plags: List[StudentInfo] = []
            course_user: CourseUser
            common_id_number_to_course_user: Dict[CommonIDNumberT, CourseUser] = {}
            for course_user in course_users:
                common_id_number: CommonIDNumberT = "(invalid)"
                try:
                    if (
                        maybe_common_id_number := course_user.user.get_common_id_number()
                    ) is None:
                        raise ValueError("Unexpected faculty")
                    common_id_number = maybe_common_id_number
                except ValueError:
                    SLACK_NOTIFIER.error(
                        f"Student with no common ID number is active in course: {course_user.user}",
                        tracebacks=traceback.format_exc(),
                    )
                course_users_in_plags.append(
                    StudentInfo(
                        common_id_number=common_id_number,
                        student_card_number=course_user.user.student_card_number,
                        full_name=course_user.user.full_name,
                    )
                )
                common_id_number_to_course_user[common_id_number] = course_user

            differences = DifferenceDetector(
                course_users_in_plags,
                course_users_in_lms,
                key_attr_name="common_id_number",
                value_attr_names=("student_card_number",),
            )

            def get_course_user_from_common_id_number(
                common_id_number: CommonIDNumberT,
            ) -> CourseUser:
                return common_id_number_to_course_user[common_id_number]

            def pair_to_conflict_data(
                pair: FromToPair[StudentInfo],
            ) -> StudentInfoWithConflict:
                return pair.to_item, get_course_user_from_common_id_number(
                    pair.from_item.common_id_number
                )

            course_users_only_in_plags = [
                get_course_user_from_common_id_number(pair.common_id_number)
                for pair in differences.deleted
            ]
            course_users_only_in_lms = differences.inserted
            course_users_in_both_with_conflict = [
                pair_to_conflict_data(pair) for pair in differences.updated
            ]
            course_users_in_both_without_conflict = [
                get_course_user_from_common_id_number(pair.from_item.common_id_number)
                for pair in differences.kept
            ]

        form = UploadCourseUserLmsExcelFile()
        return render(
            request,
            "course_user/lms_diff.html",
            dict(
                user_authority=user_authority,
                organization=organization,
                course=course,
                form=form,
                course_users_only_in_plags=course_users_only_in_plags,
                course_users_only_in_lms=course_users_only_in_lms,
                course_users_in_both_with_conflict=course_users_in_both_with_conflict,
                course_users_in_both_without_conflict=course_users_in_both_without_conflict,
            ),
        )

    @classmethod
    @annotate_view_endpoint(
        require_capabilities=(UserAuthorityCapabilityKeys.CAN_MANAGE_COURSE_USER,)
    )
    def _get(
        cls,
        request: HttpRequest,
        user_authority: UserAuthorityDict,
        organization: Organization,
        course: Course,
    ) -> HttpResponse:
        return cls._render_page(request, user_authority, organization, course)

    @classmethod
    @annotate_view_endpoint(
        require_capabilities=(UserAuthorityCapabilityKeys.CAN_MANAGE_COURSE_USER,)
    )
    def _post(
        cls,
        request: HttpRequest,
        user_authority: UserAuthorityDict,
        organization: Organization,
        course: Course,
    ) -> HttpResponse:
        form = UploadCourseUserLmsExcelFile(request.POST, request.FILES)
        if not form.is_valid():
            return cls._render_page(request, user_authority, organization, course)
        with ExceptionHandler("LMS CourseUser Diff", request):
            return cls._render_page(request, user_authority, organization, course, form)
